import openpyxl

file = openpyxl.load_workbook('Задание 3.xlsx')
sheet = file.active

# переменные с периодами
period_1 = '2023.01.01 - 2023.03.01'
period_2 = '2023.04.01 - 2023.06.01'

# Цикл для растоновки значений
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=4).value = 0

# Цикл для проверки условий
for i in range(2, sheet.max_row + 1):
    name_product = sheet.cell(row=i, column=1).value  # переменная с наименование продукта
    date = sheet.cell(row=i, column=2).value  # переменная с датой
    vg = sheet.cell(row=i, column=3).value  # переменная ВГ

    for k in range(i + 1, sheet.max_row + 1):
        name_product_2 = sheet.cell(row=k, column=1).value
        date_2 = sheet.cell(row=k, column=2).value
        vg_2 = sheet.cell(row=k, column=3).value
        # если названия совподают - значит продукт производился в 2-х периодах и далее проверяем словия
        if name_product == name_product_2 and (date == period_1) and (date_2 == period_2) and (
                vg - vg_2) > 5 and vg_2 < 90:
            sheet.cell(row=i, column=4).value = 1
            sheet.cell(row=k, column=4).value = 1

        elif name_product == name_product_2 and (date == period_2) and (date_2 == period_1) and (
                vg_2 - vg) > 5 and vg < 90:
            sheet.cell(row=i, column=4).value = 1
            sheet.cell(row=k, column=4).value = 1
        # уловие если продукты производился в одном периоде
        elif (date == period_1 or date == period_2) and (
                date_2 == period_2 or date_2 == period_1) and vg < 90 and vg_2 < 90:
            sheet.cell(row=i, column=4).value = 1
            sheet.cell(row=k, column=4).value = 1
# сохранение изменений
file.save('Задание 3.xlsx')
